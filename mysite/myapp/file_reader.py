import csv
import openpyxl
import os
from .models import studentProfile, testQuestion, studentMark

root_path = os.path.dirname(os.path.abspath(__file__))

def file_to_db(filename, client_name, test_id):
               
    if filename[-4:] == 'xlsx' or filename[-3:] == 'xls':
        data = read_xl(filename)
        write_db(data, client_name, test_id)
    else:
        print('error')

    
def read_xl(filename):
    wb = openpyxl.load_workbook(root_path + '/uploads/' + filename)
    anotherSheet = wb.active
    ws = wb.worksheets[0]
    data = {}
    for i in wb.worksheets:
        row_count = i.max_row +1 
        col_count = i.max_column
        for row in range(2,row_count):
            data[row-1] = {}
            for col in range(1,col_count):
                a = i.cell(column=col, row=row).value
                data[row-1][col-1] = a if a != None else 'na'
    data['filename'] = filename
    return data

def write_db(data, client_name, file_id):
    for i in data:
        if str(i) == 'filename':
            break
        ques=testQuestion.objects.create(
            question_id=file_id,
            question=data[i].get(1),
            option1=data[i].get(2),
            option2=data[i].get(3),
            option3=data[i].get(4),
            option4=data[i].get(5),
            answer=data[i].get(6),
        )
   

